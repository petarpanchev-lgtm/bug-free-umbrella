"""
daily_screen.py

Standalone daily scan (NOT a Streamlit app -- meant to run headless via
GitHub Actions, see .github/workflows/daily_screener.yml). Runs all 3
Kullamagi setup screeners -- Breakout, Episodic Pivot, Parabolic Short --
against the full NYSE + NASDAQ universe, using the exact same book-only
logic and default thresholds as the screener tabs in app.py.

Rather than re-writing the whole ticker list every day, this keeps a
running log in screener_history.xlsx (one sheet per setup) and only
records what CHANGED since the previous run:
  - First run ever for a setup: every current hit is logged with
    Status = "Initial".
  - Every run after that: only newly-flagged tickers ("Added") and
    tickers that no longer qualify ("Dropped") get a new row. Tickers
    that are still flagged and were already flagged yesterday get no
    row at all -- that's the point of tracking changes instead of state.

"Yesterday's" set of currently-flagged tickers is not stored separately;
it's reconstructed by replaying each sheet's own history top to bottom
(Initial/Added -> add to the set, Dropped -> remove from the set). This
keeps the workbook itself as the single source of truth, so there's
nothing else to keep in sync.

Every "Initial"/"Added" row also runs that setup's own trade-planner
calculator (the same one behind the app's Step 2 calculators) against
the ticker and logs Entry / Stop Loss / Take Profit (2R target), so the
sheet doubles as a ready-to-use trade plan, not just a candidate list.
"Dropped" rows leave those columns blank -- the setup no longer applies,
so there's no trade to plan. The calculator uses a generic $10,000 /
0.5%-risk placeholder purely to derive entry/stop/targets, which don't
actually depend on account size -- re-run the app's own calculator with
your real account size for the position size that matches your risk.

Usage:
    python daily_screen.py

Reads/writes:
    screener_history.xlsx  (created if it doesn't exist yet)
"""

import datetime as dt
import os
import sys

from openpyxl import Workbook, load_workbook

from screener import (
    get_sp500_tickers,
    get_nasdaq100_tickers,
    get_all_us_tickers,
    get_common_stocks_from_csv,
    batch_download,
)
from kullamagi_setups import (
    screen_breakouts, screen_episodic_pivots, screen_parabolic_shorts,
    calc_breakout_trade, calc_episodic_pivot_trade, calc_parabolic_short_trade,
    fetch_intraday_5m,
)

HISTORY_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "screener_history.xlsx")

SETUPS = ["Breakout", "EP", "Parabolic Short"]

HEADER = ["Date", "Ticker", "Status", "Entry", "Stop Loss", "Take Profit", "Shares (10k Acct)"]

# Placeholder inputs for the trade-planner calculators below. Entry, Stop
# Loss, and the R-multiple targets are derived purely from price data --
# account size and risk % only affect position size (Shares), which is why
# "Shares (10k Acct)" is explicitly labeled with the $10,000 / 0.5%-risk
# assumption baked into it -- re-run the app's own calculator with your real
# account size and risk % for your actual position size.
_CALC_ACCOUNT_SIZE = 10000.0
_CALC_RISK_PCT = 0.5


def _trade_params_breakout(ticker, price_data):
    """Runs calc_breakout_trade for one ticker and extracts Entry/Stop
    Loss/Take Profit (2R target)/Shares (at the $10k/0.5%-risk placeholder).
    Returns None if there's not enough data or the calculation fails for
    any reason -- caller logs blanks."""
    df = price_data.get(ticker)
    if df is None or df.empty:
        return None
    try:
        res = calc_breakout_trade(df, _CALC_ACCOUNT_SIZE, _CALC_RISK_PCT)
    except Exception:
        return None
    targets = res.get("targets") or {}
    return {
        "Entry": res.get("entry"), "Stop Loss": res.get("stop"),
        "Take Profit": targets.get("2R"), "Shares": res.get("shares"),
    }


def _trade_params_ep(ticker, price_data):
    """Runs calc_episodic_pivot_trade for one ticker. Fetches real intraday
    5-minute bars for the entry/stop if available (same as the app's
    calculator); falls back to daily open/low if not."""
    df = price_data.get(ticker)
    if df is None or df.empty:
        return None
    try:
        intraday = fetch_intraday_5m(ticker)
    except Exception:
        intraday = None
    try:
        res = calc_episodic_pivot_trade(df, intraday, _CALC_ACCOUNT_SIZE, _CALC_RISK_PCT)
    except Exception:
        return None
    targets = res.get("targets") or {}
    return {
        "Entry": res.get("entry"), "Stop Loss": res.get("stop"),
        "Take Profit": targets.get("2R"), "Shares": res.get("shares"),
    }


def _trade_params_parabolic_short(ticker, price_data):
    """Runs calc_parabolic_short_trade for one ticker. Fetches real intraday
    5-minute bars for the entry/stop if available; falls back to daily
    low/high if not."""
    df = price_data.get(ticker)
    if df is None or df.empty:
        return None
    try:
        intraday = fetch_intraday_5m(ticker)
    except Exception:
        intraday = None
    try:
        res = calc_parabolic_short_trade(df, intraday, _CALC_ACCOUNT_SIZE, _CALC_RISK_PCT)
    except Exception:
        return None
    targets = res.get("targets") or {}
    return {
        "Entry": res.get("entry"), "Stop Loss": res.get("stop"),
        "Take Profit": targets.get("2R"), "Shares": res.get("shares"),
    }


def get_universe():
    """Full NYSE+NASDAQ universe, with a two-stage fallback if the live
    symbol-directory fetch fails for any reason:
      1. Live fetch from Nasdaq Trader (get_all_us_tickers) -- freshest,
         includes brand-new listings, but depends on that site being up.
      2. Local nasdaq_nyse_common_stock.csv (get_common_stocks_from_csv) --
         no network call, kept fresh weekly by
         .github/workflows/update_common_stock_list.yml, so it's a much
         bigger and more current safety net than the old S&P 500 +
         Nasdaq-100 fallback.
      3. S&P 500 + Nasdaq-100 combined -- final safety net if even the
         bundled CSV is missing or unreadable.
    """
    tickers = get_all_us_tickers()
    if not tickers:
        print("Full universe fetch failed -- falling back to bundled common-stock CSV.", file=sys.stderr)
        tickers = get_common_stocks_from_csv()
    if not tickers:
        print("Bundled CSV fallback unavailable -- falling back to S&P 500 + Nasdaq-100 combined.", file=sys.stderr)
        tickers = list(dict.fromkeys(get_sp500_tickers() + get_nasdaq100_tickers()))
    return tickers


def reconstruct_current_state(ws):
    """Replay a history sheet's rows in date order to reconstruct the set
    of tickers considered 'currently hit' as of the last recorded run."""
    state = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        _, ticker, status = row[0], row[1], row[2]
        if ticker is None or status is None:
            continue
        if status in ("Initial", "Added"):
            state.add(ticker)
        elif status == "Dropped":
            state.discard(ticker)
    return state


def reconstruct_current_state_detailed(ws):
    """Like reconstruct_current_state, but keeps each currently-active
    ticker's most recently logged Date/Entry/Stop Loss/Take Profit/Shares
    instead of just set membership. Used by the app (app.py) to render a
    live view of screener_history.xlsx without duplicating the replay logic.

    Returns {ticker: {"Date Flagged": ..., "Entry": ..., "Stop Loss": ...,
    "Take Profit": ..., "Shares": ...}} for every ticker currently
    considered hit.
    """
    state = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        date, ticker, status = row[0], row[1], row[2]
        entry = row[3] if len(row) > 3 else None
        stop = row[4] if len(row) > 4 else None
        take_profit = row[5] if len(row) > 5 else None
        shares = row[6] if len(row) > 6 else None
        if ticker is None or status is None:
            continue
        if status in ("Initial", "Added"):
            state[ticker] = {
                "Date Flagged": date,
                "Entry": entry,
                "Stop Loss": stop,
                "Take Profit": take_profit,
                "Shares": shares,
            }
        elif status == "Dropped":
            state.pop(ticker, None)
    return state


def _find_last_active_rows(ws):
    """Replays a sheet's history like reconstruct_current_state, but returns
    {ticker: row_index} pointing at the specific Initial/Added row still
    "active" for each currently-flagged ticker -- i.e. the exact row
    update_sheet's backfill step should edit in place if that ticker's
    trade params are blank."""
    last_row = {}
    for row_idx in range(2, ws.max_row + 1):
        row = [c.value for c in ws[row_idx]]
        if not row or row[0] is None:
            continue
        ticker, status = row[1], row[2]
        if ticker is None or status is None:
            continue
        if status in ("Initial", "Added"):
            last_row[ticker] = row_idx
        elif status == "Dropped":
            last_row.pop(ticker, None)
    return last_row


def load_or_create_workbook():
    if os.path.exists(HISTORY_PATH):
        wb = load_workbook(HISTORY_PATH)
    else:
        wb = Workbook()
        wb.remove(wb.active)  # drop the default blank sheet
    for name in SETUPS:
        if name not in wb.sheetnames:
            ws = wb.create_sheet(name)
            ws.append(HEADER)
        else:
            # Backward-compat: workbooks from before the Entry/Stop Loss/Take
            # Profit columns existed only have "Date, Ticker, Status". Extend
            # the header in place rather than rewriting the sheet -- existing
            # rows just get blank cells for the new columns; only rows
            # appended from now on populate them.
            ws = wb[name]
            for col_idx, col_name in enumerate(HEADER, start=1):
                if ws.cell(row=1, column=col_idx).value != col_name:
                    ws.cell(row=1, column=col_idx, value=col_name)
    return wb


def update_sheet(wb, setup_name, today_hits, today_str, trade_params_fn, price_data):
    """Diff today's hits against the reconstructed prior state for this
    setup's sheet, and append only the rows that changed. Every "Initial" or
    "Added" row also gets that setup's calculator run against it (Entry,
    Stop Loss, Take Profit = 2R target, Shares at the $10k/0.5%-risk
    placeholder); "Dropped" rows leave those columns blank since the setup
    no longer applies.

    Tickers that are still flagged today AND were already flagged before
    ("unchanged") don't get a new row -- but if their most recent row is
    missing any of Entry/Stop Loss/Take Profit/Shares (e.g. logged by a run
    from before these columns existed), this backfills those specific cells
    in place. Otherwise a ticker flagged before this feature existed would
    stay blank forever, since it may never naturally cycle through
    Added/Dropped again.

    Returns a small summary dict for the run log."""
    ws = wb[setup_name]
    previous_state = reconstruct_current_state(ws)
    today_set = set(today_hits)

    added = sorted(today_set - previous_state)
    dropped = sorted(previous_state - today_set)
    unchanged = sorted(today_set & previous_state)
    is_first_run = ws.max_row <= 1  # only the header row present so far

    def _row(ticker, status):
        if status == "Dropped":
            params = None
        else:
            params = trade_params_fn(ticker, price_data)
        entry = params.get("Entry") if params else None
        stop = params.get("Stop Loss") if params else None
        take_profit = params.get("Take Profit") if params else None
        shares = params.get("Shares") if params else None
        return [today_str, ticker, status, entry, stop, take_profit, shares]

    backfilled = 0
    if is_first_run:
        for t in sorted(today_set):
            ws.append(_row(t, "Initial"))
        summary = {"mode": "initial", "count": len(today_set)}
    else:
        for t in added:
            ws.append(_row(t, "Added"))
        for t in dropped:
            ws.append(_row(t, "Dropped"))

        if unchanged:
            last_row_for_ticker = _find_last_active_rows(ws)
            for t in unchanged:
                row_idx = last_row_for_ticker.get(t)
                if row_idx is None:
                    continue
                existing = [ws.cell(row=row_idx, column=c).value for c in (4, 5, 6, 7)]
                if all(v is not None for v in existing):
                    continue  # already fully populated, nothing to backfill
                params = trade_params_fn(t, price_data)
                if not params:
                    continue
                ws.cell(row=row_idx, column=4, value=params.get("Entry"))
                ws.cell(row=row_idx, column=5, value=params.get("Stop Loss"))
                ws.cell(row=row_idx, column=6, value=params.get("Take Profit"))
                ws.cell(row=row_idx, column=7, value=params.get("Shares"))
                backfilled += 1

        summary = {
            "mode": "diff",
            "added": len(added),
            "dropped": len(dropped),
            "unchanged": len(unchanged),
            "backfilled": backfilled,
        }
    return summary


def main():
    today_str = dt.date.today().isoformat()
    print(f"=== Daily Kullamagi screener run: {today_str} ===")

    tickers = get_universe()
    print(f"Universe size: {len(tickers)} tickers")

    print("Downloading price data (shared across all 3 screeners)...")
    price_data = batch_download(tickers)
    print(f"Downloaded data for {sum(1 for v in price_data.values() if v is not None)} of {len(tickers)} tickers.")

    bo_hits, bo_errors = screen_breakouts(price_data)
    ep_hits, ep_errors = screen_episodic_pivots(price_data)
    ps_hits, ps_errors = screen_parabolic_shorts(price_data)

    print(f"Breakout: {len(bo_hits)} hits, {len(bo_errors)} skipped (no/insufficient data)")
    print(f"EP: {len(ep_hits)} hits, {len(ep_errors)} skipped (no/insufficient data)")
    print(f"Parabolic Short: {len(ps_hits)} hits, {len(ps_errors)} skipped (no/insufficient data)")

    wb = load_or_create_workbook()
    results = {
        "Breakout": update_sheet(
            wb, "Breakout", [h["Ticker"] for h in bo_hits], today_str,
            _trade_params_breakout, price_data,
        ),
        "EP": update_sheet(
            wb, "EP", [h["Ticker"] for h in ep_hits], today_str,
            _trade_params_ep, price_data,
        ),
        "Parabolic Short": update_sheet(
            wb, "Parabolic Short", [h["Ticker"] for h in ps_hits], today_str,
            _trade_params_parabolic_short, price_data,
        ),
    }

    wb.save(HISTORY_PATH)
    print(f"Saved {HISTORY_PATH}")

    for setup_name, s in results.items():
        if s["mode"] == "initial":
            print(f"[{setup_name}] First run recorded: {s['count']} initial hits.")
        else:
            print(f"[{setup_name}] {s['added']} added, {s['dropped']} dropped, "
                  f"{s['unchanged']} unchanged since last run "
                  f"({s.get('backfilled', 0)} unchanged tickers had blank trade params backfilled).")


if __name__ == "__main__":
    main()
