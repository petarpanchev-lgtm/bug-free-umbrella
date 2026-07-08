"""
Kullamagi Setup Tools - Streamlit web app

Six tools built ONLY from what Kristjan Kullamagi says in his own words in
his interview in Jack Schwager & George F. Coyle's "Market Wizards: The
Next Generation" (Chapter 1):

  Screeners (scan a universe, flag tickers matching ALL book conditions):
    - Breakout / Momentum Burst Screener
    - Episodic Pivot (EP) Screener
    - Parabolic Short Screener

  Calculators (type a ticker, get entry / stop / position size / targets):
    - Breakout Trade Planner
    - Episodic Pivot Trade Planner
    - Parabolic Short Trade Planner

Laid out as a two-step workflow: Step 1 picks a strategy and screens a
universe for candidates; Step 2 takes a ticker (from Step 1, or one you
already have) and calculates the exact entry/stop/position size.

See kullamagi_setups.py for the exact book citations behind every
threshold used here. Deploy on Streamlit Community Cloud (see README.md).
"""

import os
import re

import requests
import streamlit as st
import pandas as pd
from openpyxl import load_workbook

from kullamagi_score import fetch_data, market_regime
from screener import (
    get_sp500_tickers,
    get_nasdaq100_tickers,
    get_common_stocks_from_csv,
    batch_download,
)
from kullamagi_setups import (
    screen_breakouts, screen_episodic_pivots, screen_parabolic_shorts,
    calc_breakout_trade, calc_episodic_pivot_trade, calc_parabolic_short_trade,
    fetch_intraday_5m,
)
from daily_screen import HISTORY_PATH, SETUPS, reconstruct_current_state_detailed

st.set_page_config(page_title="Kullamagi Setup Tools", page_icon="📈", layout="wide")

st.title("📈 Kullamagi Setup Tools")
st.caption(
    "Screeners and trade planners built ONLY from Kristjan Kullamagi's own words in "
    "Schwager & Coyle's *Market Wizards: The Next Generation* (Chapter 1). "
    "Independent, educational tool -- not affiliated with or endorsed by him, and not financial advice."
)

with st.expander("Book citations behind these tools", expanded=False):
    st.markdown(
        """
**Breakout / "Momentum Burst"**
- Leading stock: *"scan for the 1% or 2% stocks with the largest upmoves in the past one, three, and six months."*
- Consolidation: *"I usually look for shorter consolidations -- about one to three weeks."*
- Stop: *"The stop would be at the low of the day... the stop should be no wider than the average daily true range (ADTR)... If the low is more than the ADTR below the current price, then I wouldn't take the trade."*
- Exit: *"I will take a partial profit during the first three to five days... then use a close below either a 10-day or 20-day moving average as a trailing stop."*
- Market filter (applies to Breakout & EP, not Parabolic Short): *"I use the 10-day and 20-day moving averages. When they are both moving up, the market is in an uptrend... When the 10-day crosses below the 20-day, that's a sign of caution."*

**Episodic Pivot**
- *"I look for stocks up at least 10% on high volume."*
- *"...suddenly, it has 10 times its average daily volume."*
- *"The best episodic pivot trade is a neglected stock... going sideways for a long time -- months, even years."*
- Entry: *"I will buy on a move above the high of the first 5-minute bar. Sometimes, I need to wait longer... and I will buy on a move above the first 1-hour price bar."*
- Stop: *"The low of the day."*

**Parabolic Short**
- A+ setup: *"the stock is up three or four days in a row with a total gain of at least 300%... I never go short on day one... I rarely short on day two."*
- Entry: *"The short trigger could be breaking the opening range lows or... breaking below a 5-minute candle bar."*
- Stop: *"Would your stop be at the high of the day? Yes, most of the time."*
- Market-agnostic: *"For the parabolic short trade, it doesn't matter what kind of market you are in."*

**Risk / position sizing (all setups)**
- *"Typically, I will risk 0.5% or less of my account size per trade, but I may risk up to a maximum of 1% on some trades."*
        """
    )

def trigger_github_workflow():
    """
    Manually triggers the daily_screener.yml GitHub Actions workflow via the
    GitHub REST API (a workflow_dispatch event), instead of waiting for its
    scheduled 21:30 UTC run.

    Requires three Streamlit secrets to be set (app -> Settings -> Secrets on
    Streamlit Community Cloud -- see README.md for how to create the token):
      GITHUB_TOKEN       - a fine-grained PAT scoped to this repo with
                            "Actions: Read and write" permission
      GITHUB_REPO_OWNER  - your GitHub username or org
      GITHUB_REPO_NAME   - the repo name (e.g. "kullamagi-tools")
    Optional:
      GITHUB_BRANCH      - defaults to "main"

    Returns (success: bool, message: str). Never logs or displays the token
    itself.
    """
    token = st.secrets.get("GITHUB_TOKEN")
    owner = st.secrets.get("GITHUB_REPO_OWNER")
    repo = st.secrets.get("GITHUB_REPO_NAME")
    branch = st.secrets.get("GITHUB_BRANCH", "main")

    missing = [
        name for name, val in
        [("GITHUB_TOKEN", token), ("GITHUB_REPO_OWNER", owner), ("GITHUB_REPO_NAME", repo)]
        if not val
    ]
    if missing:
        return False, (
            "Missing Streamlit secret(s): " + ", ".join(missing) + ". Add them under "
            "this app's Settings -> Secrets on Streamlit Community Cloud (see README.md "
            "for how to create the GitHub token)."
        )

    url = f"https://api.github.com/repos/{owner}/{repo}/actions/workflows/daily_screener.yml/dispatches"
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }
    try:
        resp = requests.post(url, headers=headers, json={"ref": branch}, timeout=15)
    except Exception as e:
        return False, f"Request to GitHub failed: {e}"

    if resp.status_code == 204:
        return True, (
            "Workflow run triggered! It usually takes a few seconds to appear -- check "
            f"the Actions tab at https://github.com/{owner}/{repo}/actions for progress."
        )
    elif resp.status_code == 404:
        return False, (
            "GitHub returned 404 -- double-check GITHUB_REPO_OWNER/GITHUB_REPO_NAME, that "
            "daily_screener.yml is pushed to the branch in GITHUB_BRANCH, and that the "
            "token has access to this repo."
        )
    elif resp.status_code in (401, 403):
        return False, (
            f"GitHub returned {resp.status_code} -- the token is invalid, expired, or "
            "doesn't have 'Actions: write' permission on this repo."
        )
    else:
        return False, f"GitHub API error {resp.status_code}: {resp.text[:300]}"


with st.expander("Manual trigger: run the daily screener now (GitHub Actions)", expanded=False):
    st.caption(
        "The daily scan normally runs automatically on GitHub's schedule (see "
        "daily_screener.yml, 21:30 UTC on weekdays). Use this button to kick off an "
        "extra run right now instead of waiting -- handy right after deploying, or "
        "whenever you want a fresh scan. Requires GITHUB_TOKEN, GITHUB_REPO_OWNER, and "
        "GITHUB_REPO_NAME to be set in this app's secrets (see README.md)."
    )
    if st.button("▶️ Run workflow now", type="primary"):
        with st.spinner("Triggering GitHub Actions workflow..."):
            ok, msg = trigger_github_workflow()
        if ok:
            st.success(msg)
        else:
            st.error(msg)


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------
# (render_daily_scan_results is defined here; called further down, after all
# helper/render functions are defined, in the page-layout section.)

def render_daily_scan_results():
    """
    Reads screener_history.xlsx straight from disk (the same file
    daily_screen.py writes and GitHub Actions commits back to the repo on
    schedule) and shows, per setup, the tickers currently flagged along with
    the Entry/Stop Loss/Take Profit/Shares logged when each was first
    flagged -- plus the full Initial/Added/Dropped history underneath.
    Shares assumes a $10,000 account and 0.5% risk per trade (see
    daily_screen.py's _CALC_ACCOUNT_SIZE/_CALC_RISK_PCT) -- Entry/Stop
    Loss/Take Profit don't depend on account size, but Shares does, so
    re-run the app's own calculator below with your real numbers for your
    actual position size.

    No extra infrastructure needed: Streamlit Community Cloud auto-redeploys
    on every push to the repo, including the automated daily commit, so this
    view stays in sync with at most a couple of minutes' lag after each run.

    Also self-heals: if a ticker is still flagged today but its logged row
    predates these columns (blank Entry/Stop Loss/Take Profit/Shares), the
    next daily_screen.py run backfills those cells in place instead of
    leaving them blank forever.
    """
    if not os.path.exists(HISTORY_PATH):
        st.info(
            "No automated scan history yet -- screener_history.xlsx is created after the "
            "first run of the daily GitHub Actions screener. Trigger it manually above, or "
            "wait for the next scheduled run (weekdays, 21:30 UTC)."
        )
        return

    try:
        wb = load_workbook(HISTORY_PATH, data_only=True)
    except Exception as e:
        st.error(f"Couldn't read screener_history.xlsx: {e}")
        return

    last_run_dates = []
    for setup_name in SETUPS:
        if setup_name not in wb.sheetnames:
            continue
        ws = wb[setup_name]

        all_rows = list(ws.iter_rows(min_row=1, values_only=True))
        header, data_rows = (all_rows[0], all_rows[1:]) if all_rows else ([], [])
        dates_in_sheet = [r[0] for r in data_rows if r and r[0]]
        if dates_in_sheet:
            last_run_dates.append(max(str(d) for d in dates_in_sheet))

        state = reconstruct_current_state_detailed(ws)
        st.subheader(f"{setup_name} -- currently flagged ({len(state)})")
        if not state:
            st.caption("No tickers currently flagged for this setup.")
        else:
            rows = [{"Ticker": t, **v} for t, v in state.items()]
            rows.sort(key=lambda r: str(r.get("Date Flagged") or ""), reverse=True)
            st.dataframe(pd.DataFrame(rows), hide_index=True, width="stretch")

        with st.expander(f"{setup_name}: full history (all Initial/Added/Dropped rows)"):
            if data_rows:
                st.dataframe(pd.DataFrame(data_rows, columns=header), hide_index=True, width="stretch")
            else:
                st.caption("No history rows yet.")

    if last_run_dates:
        st.caption(f"Last recorded run: {max(last_run_dates)}")

    with open(HISTORY_PATH, "rb") as f:
        st.download_button(
            "Download full screener_history.xlsx", f.read(),
            "screener_history.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_history_xlsx",
        )


def render_market_environment(me):
    if not me or me.get("trend") == "unknown":
        return
    if me["trend"] == "uptrend":
        st.success(
            f"Market environment: **favorable** (SPY 10-day MA {me['sma10']} is above its "
            f"20-day MA {me['sma20']}) -- book says only take Breakout/EP setups here."
        )
    else:
        st.error(
            f"Market environment: **unfavorable** (SPY 10-day MA {me['sma10']} is below its "
            f"20-day MA {me['sma20']}) -- book says stand aside on Breakout/EP setups here. "
            "(Parabolic Short is unaffected by market direction.)"
        )


def universe_picker(key_prefix):
    """Renders universe-selection widgets + a run button. Returns the
    ticker list only on the run the button was clicked, else None."""
    universe_choice = st.radio(
        "Universe",
        [
            "S&P 500",
            "Nasdaq-100",
            "All NYSE and NASDAQ common stocks (3963 tickers)",
            "Custom list",
        ],
        horizontal=True, key=f"{key_prefix}_universe",
        help="'All NYSE and NASDAQ common stocks' reads a local CSV (no ETFs) "
             "refreshed weekly -- no live symbol-directory fetch needed, so it's "
             "fast and reliable. Ticker count drifts slightly as that weekly "
             "refresh runs (new listings/delistings).",
    )
    custom_tickers = None
    if universe_choice == "Custom list":
        custom_input = st.text_area(
            "Paste tickers (comma, space, or newline separated)",
            height=80, key=f"{key_prefix}_custom_text",
        )
        uploaded = st.file_uploader(
            "...or upload a CSV with a Ticker/Symbol column",
            type=["csv"], key=f"{key_prefix}_custom_csv",
        )
        if uploaded is not None:
            try:
                cdf = pd.read_csv(uploaded)
                col = next(
                    (c for c in cdf.columns if str(c).strip().lower() in ("ticker", "tickers", "symbol", "symbols")),
                    cdf.columns[0],
                )
                custom_tickers = cdf[col].astype(str).tolist()
            except Exception as e:
                st.error(f"Couldn't read that CSV: {e}")
        elif custom_input.strip():
            custom_tickers = [t for t in re.split(r"[,\s]+", custom_input.strip()) if t]

    max_tickers = st.number_input(
        "Max tickers to scan (safety cap)", min_value=10, max_value=10000,
        value=3000, step=100, key=f"{key_prefix}_max",
        help="Raise this to scan the full 'All NYSE and NASDAQ common stocks' "
             "universe (~4,000 tickers). Larger scans take longer -- expect "
             "several minutes for 3,000+, and Yahoo Finance may rate-limit very "
             "large scans (the app retries failed batches automatically, but if "
             "you keep seeing rate-limit errors, lower this number or try again "
             "in a few minutes).",
    )

    if st.button("Fetch universe & run", type="primary", width="stretch", key=f"{key_prefix}_run"):
        if universe_choice == "S&P 500":
            with st.spinner("Fetching S&P 500 list..."):
                tickers = get_sp500_tickers()
        elif universe_choice == "Nasdaq-100":
            with st.spinner("Fetching Nasdaq-100 list..."):
                tickers = get_nasdaq100_tickers()
        elif universe_choice == "All NYSE and NASDAQ common stocks (3963 tickers)":
            with st.spinner("Loading bundled common-stock list..."):
                tickers = get_common_stocks_from_csv()
                if not tickers:
                    st.error(
                        "Couldn't read the bundled common-stock CSV (file missing or "
                        "unreadable). Falling back to S&P 500 + Nasdaq-100 combined."
                    )
                    tickers = list(dict.fromkeys(get_sp500_tickers() + get_nasdaq100_tickers()))
        else:
            tickers = custom_tickers or []
        # Belt-and-suspenders: whatever the source, make sure every entry
        # reaching the downloader/screeners is a clean, non-empty string.
        # (A malformed row from a live feed like the Nasdaq Trader symbol
        # directory once slipped a bad value through and crashed the
        # skipped-tickers display -- this stops that class of bug at the
        # source instead of relying on every downstream consumer to guard.)
        tickers = [str(t).strip() for t in (tickers or []) if str(t).strip()]
        return tickers[: int(max_tickers)] if tickers else []
    return None


def fetch_batch_with_progress(tickers):
    progress = st.progress(0.0)
    status = st.empty()

    def _cb(done, total):
        progress.progress(done / total)
        status.text(f"{done}/{total} tickers fetched...")

    price_data = batch_download(tickers, progress_cb=_cb)
    progress.empty()
    status.empty()
    return price_data


def render_calc_result(res, account_size):
    if res.get("risk_per_share") is None or res.get("shares", 0) <= 0:
        st.error(
            "Risk per share is zero or negative (entry/stop don't make sense here) -- "
            "can't size a position. Check the ticker's recent data."
        )

    if "stop_within_adtr" in res and not res["stop_within_adtr"]:
        st.warning(
            f"Stop distance (${res['risk_per_share']}) is WIDER than 1x ADTR (${res['adtr']}). "
            "Per the book's rule, this means: **skip the trade.**"
        )

    if "gap_ok" in res and not res["gap_ok"]:
        st.warning(f"Gap is only {res['gap_pct']}% -- book wants at least 10%.")
    if "volume_ok" in res and not res["volume_ok"]:
        st.warning(f"Volume is only {res['volume_multiple']}x the 50-day average -- below your threshold.")
    if "a_plus_setup" in res and not res["a_plus_setup"]:
        st.warning(
            f"Consecutive up days: {res['consecutive_up_days']}, cumulative gain: "
            f"{res['cumulative_gain_pct']}% -- doesn't clear the book's A+ bar (3-5 days, 300%+)."
        )

    c1, c2, c3 = st.columns(3)
    c1.metric("Entry", f"${res['entry']}")
    c2.metric("Stop", f"${res['stop']}")
    c3.metric("Risk / share", f"${res.get('risk_per_share', '—')}")

    c4, c5, c6 = st.columns(3)
    c4.metric("Shares", f"{res.get('shares', 0):,}")
    c5.metric("Position value", f"${res.get('position_value', 0):,.0f}")
    c6.metric("% of account", f"{res.get('position_pct_of_account', '—')}%")

    if res.get("targets"):
        t = res["targets"]
        c7, c8, c9 = st.columns(3)
        c7.metric("1R target", f"${t.get('1R', '—')}")
        c8.metric("2R target", f"${t.get('2R', '—')}")
        c9.metric("3R target", f"${t.get('3R', '—')}")

    if res.get("used_intraday_data") is False:
        st.info("No live intraday data available -- entry/stop above are daily-bar approximations. "
                 "Confirm against the real opening range before trading.")

    if res.get("note"):
        st.caption(res["note"])


# ---------------------------------------------------------------------------
# Screeners (Step 1 renderers)
# ---------------------------------------------------------------------------

def render_breakout_screener():
    st.subheader("Breakout / Momentum Burst Screener")
    st.caption(
        "Flags tickers that are (1) a leading stock -- top 1-2% of THIS batch by 1/3/6-month "
        "return, (2) in a tight ~1-3 week consolidation surfing the 10/20-day MA, and (3) have a "
        "stop (base low) that doesn't exceed 1x ADTR."
    )
    c1, c2, c3 = st.columns(3)
    bo_top_pct = c1.slider("Leading-stock top %", 1.0, 10.0, 2.0, 0.5, key="bo_top_pct")
    bo_cons_days = c2.slider("Consolidation window (trading days)", 5, 15, 10, 1, key="bo_cons_days")
    bo_near_pct = c3.slider("Flag 'near trigger' within %", 1.0, 15.0, 5.0, 0.5, key="bo_near_pct")

    bo_tickers = universe_picker("bo_scr")
    if bo_tickers is not None:
        if not bo_tickers:
            st.warning("No tickers to scan.")
        else:
            st.info(f"Scanning {len(bo_tickers)} tickers...")
            price_data = fetch_batch_with_progress(bo_tickers)

            spy_df = None
            try:
                spy_df = fetch_data("SPY")
            except Exception:
                pass
            render_market_environment(market_regime(spy_df))

            hits, errors = screen_breakouts(
                price_data, top_pct=bo_top_pct, consolidation_days=bo_cons_days, near_trigger_pct=bo_near_pct
            )
            if not hits:
                st.warning("No tickers matched all three Breakout conditions.")
            else:
                st.success(f"{len(hits)} of {len(bo_tickers)} tickers match the Breakout setup.")
                df_hits = pd.DataFrame(hits)
                st.dataframe(df_hits, hide_index=True, width="stretch")
                st.download_button(
                    "Download CSV", df_hits.to_csv(index=False).encode("utf-8"),
                    "breakout_screener.csv", "text/csv", key="bo_dl",
                )
            if errors:
                with st.expander(f"{len(errors)} tickers skipped (no/insufficient data)"):
                    st.write(", ".join(str(e) for e in errors))


def render_ep_screener():
    st.subheader("Episodic Pivot Screener")
    st.caption(
        "Flags tickers gapping up at least the threshold below on volume well above their "
        "50-day average. A genuine news catalyst is REQUIRED per the book but can't be verified "
        "from price data -- always confirm one manually before trading anything shown here."
    )
    c1, c2, c3 = st.columns(3)
    ep_gap = c1.slider("Minimum gap % at the open", 5.0, 50.0, 10.0, 1.0, key="ep_gap")
    ep_vol = c2.slider("Minimum volume vs 50d avg (x)", 1.0, 15.0, 3.0, 0.5, key="ep_vol",
                        help="Book figure is ~10x for the OPENING volume specifically; this compares "
                             "full-day volume to the 50-day average, a coarser proxy, so the default "
                             "is set lower than 10x on purpose.")
    ep_dorm = c3.slider("Dormancy range % (6mo, lower = flatter)", 10.0, 100.0, 40.0, 5.0, key="ep_dorm")

    ep_tickers = universe_picker("ep_scr")
    if ep_tickers is not None:
        if not ep_tickers:
            st.warning("No tickers to scan.")
        else:
            st.info(f"Scanning {len(ep_tickers)} tickers...")
            price_data = fetch_batch_with_progress(ep_tickers)

            spy_df = None
            try:
                spy_df = fetch_data("SPY")
            except Exception:
                pass
            render_market_environment(market_regime(spy_df))

            hits, errors = screen_episodic_pivots(
                price_data, min_gap_pct=ep_gap, min_volume_multiple=ep_vol, dormancy_range_pct=ep_dorm
            )
            if not hits:
                st.warning("No tickers matched the gap + volume conditions.")
            else:
                st.success(f"{len(hits)} of {len(ep_tickers)} tickers match the EP gap/volume setup.")
                df_hits = pd.DataFrame(hits)
                st.dataframe(df_hits, hide_index=True, width="stretch")
                st.download_button(
                    "Download CSV", df_hits.to_csv(index=False).encode("utf-8"),
                    "ep_screener.csv", "text/csv", key="ep_dl",
                )
            if errors:
                with st.expander(f"{len(errors)} tickers skipped (no/insufficient data)"):
                    st.write(", ".join(str(e) for e in errors))


def render_parabolic_short_screener():
    st.subheader("Parabolic Short Screener")
    st.caption(
        "Flags tickers up 3-5 consecutive days with a cumulative gain clearing the threshold "
        "below -- the book's 'A+ setup' bar. Market-agnostic (no environment filter applies here). "
        "Still requires an intraday sign of weakness before entering -- this is a candidate list."
    )
    c1, c2, c3 = st.columns(3)
    ps_min_run = c1.slider("Minimum consecutive up days", 3, 5, 3, 1, key="ps_min_run")
    ps_max_run = c2.slider("Maximum consecutive up days", 3, 6, 5, 1, key="ps_max_run")
    ps_min_gain = c3.slider("Minimum cumulative gain % over the run", 50.0, 500.0, 300.0, 10.0, key="ps_min_gain")

    ps_tickers = universe_picker("ps_scr")
    if ps_tickers is not None:
        if not ps_tickers:
            st.warning("No tickers to scan.")
        else:
            st.info(f"Scanning {len(ps_tickers)} tickers...")
            price_data = fetch_batch_with_progress(ps_tickers)

            hits, errors = screen_parabolic_shorts(
                price_data, min_run_days=ps_min_run, max_run_days=ps_max_run, min_gain_pct=ps_min_gain
            )
            if not hits:
                st.warning("No tickers matched the A+ Parabolic Short setup.")
            else:
                st.success(f"{len(hits)} of {len(ps_tickers)} tickers match the Parabolic Short setup.")
                df_hits = pd.DataFrame(hits)
                st.dataframe(df_hits, hide_index=True, width="stretch")
                st.download_button(
                    "Download CSV", df_hits.to_csv(index=False).encode("utf-8"),
                    "parabolic_short_screener.csv", "text/csv", key="ps_dl",
                )
            if errors:
                with st.expander(f"{len(errors)} tickers skipped (no/insufficient data)"):
                    st.write(", ".join(str(e) for e in errors))


# ---------------------------------------------------------------------------
# Calculators (Step 2 renderers)
# ---------------------------------------------------------------------------

def render_breakout_calculator():
    st.subheader("Breakout Trade Planner")
    st.caption(
        "Entry = high of the recent consolidation. Stop = consolidation low (stand-in for "
        "\"the low of the day\" -- use the actual day's low once you're in the trade), capped "
        "at 1x ADTR per the book's rule."
    )
    bo_ticker = st.text_input("Ticker", key="bo_calc_ticker").strip().upper()
    c1, c2, c3 = st.columns(3)
    bo_account = c1.number_input("Account size ($)", min_value=100.0, value=10000.0, step=100.0, key="bo_calc_acct")
    bo_risk = c2.slider("Risk % per trade", 0.1, 1.0, 0.5, 0.1, key="bo_calc_risk")
    bo_cons = c3.slider("Consolidation window (days)", 5, 15, 10, 1, key="bo_calc_cons")

    if st.button("Calculate", type="primary", key="bo_calc_btn"):
        if not bo_ticker:
            st.warning("Enter a ticker.")
        else:
            with st.spinner("Fetching data..."):
                df = fetch_data(bo_ticker)
            if df is None or len(df) < bo_cons + 5:
                st.error("No usable data for that ticker.")
            else:
                res = calc_breakout_trade(df, bo_account, bo_risk, bo_cons)
                render_calc_result(res, bo_account)


def render_ep_calculator():
    st.subheader("Episodic Pivot Trade Planner")
    st.caption(
        "Entry = high of the first 5-minute bar (or the first 1-hour bar if confirming volume "
        "takes longer). Stop = the low of the day. Uses live intraday data if the market is open "
        "and data is available; otherwise falls back to today's open/low as a rough stand-in."
    )
    ep_ticker = st.text_input("Ticker", key="ep_calc_ticker").strip().upper()
    c1, c2, c3 = st.columns(3)
    ep_account = c1.number_input("Account size ($)", min_value=100.0, value=10000.0, step=100.0, key="ep_calc_acct")
    ep_risk = c2.slider("Risk % per trade", 0.1, 1.0, 0.5, 0.1, key="ep_calc_risk")
    ep_min_gap_c = c3.slider("Minimum gap % check", 5.0, 30.0, 10.0, 1.0, key="ep_calc_gap")

    if st.button("Calculate", type="primary", key="ep_calc_btn"):
        if not ep_ticker:
            st.warning("Enter a ticker.")
        else:
            with st.spinner("Fetching data..."):
                daily = fetch_data(ep_ticker)
                intraday = fetch_intraday_5m(ep_ticker)
            if daily is None or len(daily) < 55:
                st.error("No usable data for that ticker.")
            else:
                res = calc_episodic_pivot_trade(daily, intraday, ep_account, ep_risk, min_gap_pct=ep_min_gap_c)
                render_calc_result(res, ep_account)


def render_parabolic_short_calculator():
    st.subheader("Parabolic Short Trade Planner")
    st.caption(
        "Entry (short trigger) = break of the opening range low (first 5-minute bar's low), or a "
        "break below a 5-minute candle. Stop = the high of the day. Uses live intraday data if "
        "available; otherwise falls back to today's daily low/high as a rough stand-in."
    )
    ps_ticker = st.text_input("Ticker", key="ps_calc_ticker").strip().upper()
    c1, c2, c3 = st.columns(3)
    ps_account = c1.number_input("Account size ($)", min_value=100.0, value=10000.0, step=100.0, key="ps_calc_acct")
    ps_risk = c2.slider("Risk % per trade", 0.1, 1.0, 0.5, 0.1, key="ps_calc_risk")
    ps_min_gain_c = c3.slider("A+ minimum cumulative gain %", 50.0, 500.0, 300.0, 10.0, key="ps_calc_gain")

    if st.button("Calculate", type="primary", key="ps_calc_btn"):
        if not ps_ticker:
            st.warning("Enter a ticker.")
        else:
            with st.spinner("Fetching data..."):
                daily = fetch_data(ps_ticker)
                intraday = fetch_intraday_5m(ps_ticker)
            if daily is None or len(daily) < 10:
                st.error("No usable data for that ticker.")
            else:
                res = calc_parabolic_short_trade(daily, intraday, ps_account, ps_risk, min_gain_pct=ps_min_gain_c)
                render_calc_result(res, ps_account)


# ---------------------------------------------------------------------------
# Page layout: automated results, then the two-step manual workflow
#   Latest Automated Daily Scan Results -- read screener_history.xlsx as-is.
#   Row 1 (Step 1) -- pick a strategy, screen a universe for candidates.
#   Row 2 (Step 2) -- take a ticker (from Step 1, or one you already have)
#                      and calculate its exact entry/stop/position size.
# ---------------------------------------------------------------------------

st.divider()
st.header("📋 Latest Automated Daily Scan Results")
st.caption(
    "Read directly from screener_history.xlsx, updated once a day by the GitHub Actions "
    "workflow (or on-demand via the manual trigger above). Shows tickers currently flagged "
    "as of the most recent run, with the Entry/Stop Loss/Take Profit/Shares logged when each "
    "was first flagged. Shares assumes a $10,000 account and 0.5% risk per trade -- use the "
    "calculator in Step 2 below with your real account size and risk % for your actual "
    "position size."
)
render_daily_scan_results()

st.divider()
st.header("Step 1 -- Select a strategy and scan")
st.caption(
    "Pick which of the three book setups to screen for, choose a universe, and click "
    "'Fetch universe & run' to find candidates matching all of that setup's conditions."
)
screener_choice = st.radio(
    "Strategy to screen",
    ["🟢 Breakout / Momentum Burst", "🚀 Episodic Pivot (EP)", "🔻 Parabolic Short"],
    horizontal=True, key="screener_strategy_choice",
)
if screener_choice == "🟢 Breakout / Momentum Burst":
    render_breakout_screener()
elif screener_choice == "🚀 Episodic Pivot (EP)":
    render_ep_screener()
else:
    render_parabolic_short_screener()

st.divider()
st.header("Step 2 -- Calculate trade parameters")
st.caption(
    "Got a candidate ticker (from Step 1 above, or one you already have in mind)? Pick the "
    "matching strategy, type in the ticker, and click 'Calculate' for the exact entry, stop, "
    "and position size."
)
calc_choice = st.radio(
    "Strategy to calculate",
    ["🧮 Breakout", "🧮 Episodic Pivot (EP)", "🧮 Parabolic Short"],
    horizontal=True, key="calc_strategy_choice",
)
if calc_choice == "🧮 Breakout":
    render_breakout_calculator()
elif calc_choice == "🧮 Episodic Pivot (EP)":
    render_ep_calculator()
else:
    render_parabolic_short_calculator()

st.divider()
st.caption(
    "Educational tool only. Not financial advice. Not affiliated with or endorsed by Kristjan "
    "Kullamagi / Qullamaggie. Data via Yahoo Finance (yfinance)."
)
